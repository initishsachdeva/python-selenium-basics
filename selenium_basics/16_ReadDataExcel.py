import openpyxl

workBook = openpyxl.load_workbook("../excelfile.xlsx")
sheet = workBook.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

# writing data to specific columns
sheet.cell(row=2, column=2).value = "def"
print(sheet.cell(row=2, column=2).value)

# to get max rows
print(sheet.max_row)

# to get max columns
print(sheet.max_column)

print(sheet['A3'].value)

# to print every value of sheet
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)

# to print specific value of sheet
for k in range(1, sheet.max_row + 1):
    if sheet.cell(row=k, column=1).value == "TestCase":
        for l in range(1, sheet.max_column + 1):
            print(sheet.cell(row=k, column=l).value)

dict = {}
# loading excel data to dictionary

for m in range(1, sheet.max_row + 1):
    if sheet.cell(row=m, column=1).value == "TestCase":
        for n in range(2, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=n).value] = sheet.cell(row=m, column=n).value
            print(dict)
